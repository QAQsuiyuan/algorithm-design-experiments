"""将剩余所有原始数据转换为 Excel"""
import os
import re
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
RESULT_DIR = os.path.join(BASE_DIR, "results")
os.makedirs(RESULT_DIR, exist_ok=True)


# ============================================================
# 一、背包数据 → knapsack_all_data.xlsx（按 n 分组，一个 n 一个 sheet）
# ============================================================
def convert_knapsack():
    knap_dir = os.path.join(DATA_DIR, "knapsack")
    # 收集所有文件并按 n 分组
    files = [f for f in os.listdir(knap_dir) if f.endswith(".txt")]

    # 按 n 分组:  {"n1000": [(cap, filepath), ...], ...}
    groups = {}
    for f in files:
        filepath = os.path.join(knap_dir, f)
        with open(filepath, "r", encoding="utf-8") as fh:
            first_line = fh.readline().strip()
        parts = first_line.split()
        n = int(parts[0])
        cap = int(parts[1]) if len(parts) > 1 else None
        key = f"n{n}"
        if key not in groups:
            groups[key] = []
        groups[key].append((cap, filepath, first_line))

    # 排序
    def sort_key(k):
        return int(k[1:])  # "n1000" → 1000
    sorted_keys = sorted(groups.keys(), key=sort_key)

    # 测试文件单独处理
    test_files = {"test_small": None, "test_n50": None}
    for f in files:
        if f.startswith("test_"):
            test_files[f.replace(".txt", "")] = os.path.join(knap_dir, f)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for key in sorted_keys:
        entries = groups[key]  # [(cap, filepath, first_line), ...]
        n = int(key[1:])
        caps = sorted([e[0] for e in entries])

        # 读第一个文件获取物品数据（同 n 下物品相同）
        _, filepath, _ = entries[0]
        with open(filepath, "r", encoding="utf-8") as fh:
            lines = fh.readlines()
        items = [line.strip().split() for line in lines[1:]]

        sheet_name = key  # "n1000", "n2000", ...
        ws = wb.create_sheet(title=sheet_name)

        # 表头
        ws.append(["物品数量 n", n])
        ws.append(["背包容量 C", ", ".join(str(c) for c in caps)])
        ws.append([])
        ws.append(["物品编号", "重量 (weight)", "价值 (value)"])

        for idx, (w, v) in enumerate(items, start=1):
            ws.append([idx, float(w), float(v)])

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18

    # 测试数据 sheet
    for test_name, filepath in test_files.items():
        if not filepath:
            continue
        with open(filepath, "r", encoding="utf-8") as fh:
            lines = fh.readlines()
        header = lines[0].strip().split()
        n_test, cap_test = int(header[0]), int(header[1])
        items = [line.strip().split() for line in lines[1:]]

        ws = wb.create_sheet(title=test_name)
        ws.append(["物品数量 n", n_test])
        ws.append(["背包容量 C", cap_test])
        ws.append([])
        ws.append(["物品编号", "重量 (weight)", "价值 (value)"])
        for idx, (w, v) in enumerate(items, start=1):
            ws.append([idx, float(w), float(v)])
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18

    outpath = os.path.join(RESULT_DIR, "knapsack_all_data.xlsx")
    wb.save(outpath)
    print(f"[OK] 背包数据: {outpath}  ({len(sorted_keys) + len(test_files)} sheets)")


# ============================================================
# 二、排序数据 → sort_data.xlsx（一个规模一个 sheet）
# ============================================================
def convert_sort():
    sort_dir = os.path.join(DATA_DIR, "sort")
    files = sorted([f for f in os.listdir(sort_dir) if f.endswith(".txt")])

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for f in files:
        filepath = os.path.join(sort_dir, f)
        with open(filepath, "r", encoding="utf-8") as fh:
            n = int(fh.readline().strip())
            data = [int(line.strip()) for line in fh]

        # sheet 名用文件名（去掉 .txt）
        sheet_name = f.replace(".txt", "")
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet 名最长 31 字符

        ws.append(["数据规模 n", n])
        ws.append([])
        ws.append(["序号", "数值"])

        for idx, val in enumerate(data, start=1):
            ws.append([idx, val])

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 15

    outpath = os.path.join(RESULT_DIR, "sort_data.xlsx")
    wb.save(outpath)
    print(f"[OK] 排序数据: {outpath}  ({len(files)} sheets)")


# ============================================================
# 三、背包实验结果 → knapsack_results.xlsx（表格形式）
# ============================================================
def convert_knapsack_results():
    """将 knapsack_results.txt 转为标准 Excel 表格"""
    filepath = os.path.join(RESULT_DIR, "knapsack_results.txt")
    if not os.path.exists(filepath):
        print("[跳过] knapsack_results.txt 不存在")
        return

    with open(filepath, "r", encoding="utf-8") as fh:
        lines = fh.readlines()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "汇总结果"

    # 写入标题和表头
    ws.append(["0-1背包问题实验结果"])
    ws.append(["算法: DP=动态规划, GR=贪心法, BT=回溯法 (-1 表示跳过或超时)"])
    ws.append([])
    ws.append(["n", "C", "DP_value", "DP_ms", "GR_value", "GR_ms", "BT_value", "BT_ms"])

    for line in lines[4:]:  # 跳过前 4 行（标题+表头）
        parts = line.strip().split()
        if len(parts) < 8:
            continue
        try:
            row = [int(parts[0]), int(parts[1])] + [float(p) for p in parts[2:8]]
            ws.append(row)
        except ValueError:
            continue

    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 16

    outpath = os.path.join(RESULT_DIR, "knapsack_results.xlsx")
    wb.save(outpath)
    print(f"[OK] 背包结果: {outpath}")


# ============================================================
# 四、排序实验结果 → sort_results.xlsx
# ============================================================
def convert_sort_results():
    """将 sort_experiment_2.txt 和 sort_experiment_1.txt 转为 Excel"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- 实验② 规模对比 ---
    filepath = os.path.join(RESULT_DIR, "sort_experiment_2.txt")
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as fh:
            lines = fh.readlines()
        ws = wb.create_sheet(title="规模对比")
        ws.append(["排序实验②: 不同输入规模下比较次数"])
        ws.append([])
        ws.append(["n", "冒泡排序", "合并排序", "快速排序"])
        for line in lines[3:]:
            parts = line.strip().split()
            if len(parts) >= 4:
                ws.append([int(parts[0]), int(parts[1]), int(parts[2]), int(parts[3])])
        for col in ["A", "B", "C", "D"]:
            ws.column_dimensions[col].width = 18

    # --- 实验① 等价类 ---
    filepath = os.path.join(RESULT_DIR, "sort_experiment_1.txt")
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as fh:
            lines = fh.readlines()
        ws = wb.create_sheet(title="等价类分析")
        ws.append(["排序实验①: 输入数据等价类分析"])
        ws.append([])
        ws.append(["批次", "冒泡排序", "合并排序", "快速排序"])
        for line in lines:
            m = re.search(r"批次(\d): 冒泡=(\d+), 合并=(\d+), 快速=(\d+)", line)
            if m:
                ws.append([int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))])
        for col in ["A", "B", "C", "D"]:
            ws.column_dimensions[col].width = 18

    outpath = os.path.join(RESULT_DIR, "sort_results.xlsx")
    wb.save(outpath)
    print(f"[OK] 排序结果: {outpath}")


# ============================================================
if __name__ == "__main__":
    print("开始转换所有原始数据为 Excel...\n")
    convert_knapsack()
    convert_sort()
    convert_knapsack_results()
    convert_sort_results()
    print("\n全部转换完成!")
